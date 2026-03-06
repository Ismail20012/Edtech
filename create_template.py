# -*- coding: utf-8 -*-
#
# Copyright (C) 2025 Gemini Project. All rights reserved.
#
# This software is licensed under the MIT License.
# For the full license text, see the LICENSE file in the project root.
#

"""
Template Excel file creator for Moroccan grade reports.
Creates pristine template for Excel generation.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path


def create_grade_template(output_path: str = "resources/template_grade.xlsx") -> bool:
    """Create a template Excel file for grade reports."""
    try:
        # --- MOD BEGIN: Template Creation ---
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Bulletin de classe"
        
        # Header styling
        header_font = Font(bold=True, size=14, name='Arial')
        subheader_font = Font(bold=True, size=12, name='Arial')
        normal_font = Font(size=10, name='Arial')
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Main header
        ws['A1'] = "ROYAUME DU MAROC"
        ws['A1'].font = header_font
        ws.merge_cells('A1:J1')
        
        ws['A2'] = "MINISTÈRE DE L'ÉDUCATION NATIONALE"
        ws['A2'].font = subheader_font
        ws.merge_cells('A2:J2')
        
        ws['A3'] = "BULLETIN DE NOTES - {TERM_NAME}"
        ws['A3'].font = header_font
        ws.merge_cells('A3:J3')
        
        # Class info
        ws['A5'] = "Classe: {CLASS_NAME}"
        ws['A5'].font = subheader_font
        
        ws['F5'] = "Année scolaire: {ACADEMIC_YEAR}"
        ws['F5'].font = subheader_font
        
        # Headers row
        headers = ['N°', 'Nom de l\'étudiant', 'N° Étudiant']
        
        # Add subject placeholders
        for i in range(7):  # 7 subjects
            headers.append(f'{{SUBJECT_{i+1}}}\n(Coef. {{COEFF_{i+1}}})')
        
        headers.extend(['Moyenne', 'Rang', 'Observation'])
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = subheader_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            if col > 3 and col <= 10:  # Subject columns
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Sample data rows (placeholders)
        for row in range(8, 28):  # 20 students max
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.font = normal_font
                
                if col == 1:  # Row number
                    cell.value = row - 7
                elif col == 2:  # Student name
                    cell.value = f"{{STUDENT_NAME_{row-7}}}"
                elif col == 3:  # Student ID
                    cell.value = f"{{STUDENT_ID_{row-7}}}"
                elif col > 3 and col <= 10:  # Grades
                    cell.value = f"{{GRADE_{row-7}_{col-3}}}"
                    cell.alignment = Alignment(horizontal='center')
                elif col == 11:  # Average
                    cell.value = f"{{AVERAGE_{row-7}}}"
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = Alignment(horizontal='center')
                elif col == 12:  # Rank
                    cell.value = f"{{RANK_{row-7}}}"
                    cell.alignment = Alignment(horizontal='center')
                elif col == 13:  # Observation
                    cell.value = f"{{OBSERVATION_{row-7}}}"
        
        # Footer
        ws['A30'] = "Directeur:"
        ws['A30'].font = subheader_font
        
        ws['F30'] = "Professeur principal:"
        ws['F30'].font = subheader_font
        
        # Auto-adjust column widths
        column_widths = {}
        for row in ws.rows:
            for cell in row:
                if cell.coordinate not in ws.merged_cells:  # Skip merged cells
                    column_letter = cell.column_letter
                    if column_letter not in column_widths:
                        column_widths[column_letter] = 0
                    try:
                        if cell.value and len(str(cell.value)) > column_widths[column_letter]:
                            column_widths[column_letter] = len(str(cell.value))
                    except:
                        pass
        
        for column_letter, width in column_widths.items():
            adjusted_width = min(width + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save template
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        
        return True
        # --- MOD END ---
        
    except Exception as e:
        print(f"Erreur création template: {e}")
        return False


if __name__ == "__main__":
    if create_grade_template():
        print("Template créé avec succès: resources/template_grade.xlsx")
    else:
        print("Erreur lors de la création du template")

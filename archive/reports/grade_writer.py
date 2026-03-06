# -*- coding: utf-8 -*-
#
# Copyright (C) 2025 Gemini Project. All rights reserved.
#
# This software is licensed under the MIT License.
# For the full license text, see the LICENSE file in the project root.
#

"""
Excel grade report generator for Moroccan school system.
Generates comprehensive grade reports using templates.
"""

import pandas as pd
import logging
from typing import Dict, List, Optional, Tuple
from PyQt6.QtSql import QSqlQuery, QSqlDatabase
from pathlib import Path
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


class GradeReportGenerator:
    """Generates Excel grade reports with formatting."""
    
    def __init__(self, db: QSqlDatabase):
        self.db = db
        self.logger = logging.getLogger(__name__)
        
        # Report styling
        self.header_font = Font(bold=True, size=12, name='Arial')
        self.student_font = Font(size=10, name='Arial')
        self.grade_font = Font(size=10, name='Arial', bold=True)
        
        # Cell alignment
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        
        # Borders
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Grade color coding
        self.grade_colors = {
            'excellent': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),  # Light green
            'good': PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid'),      # Moccasin
            'average': PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid'),   # Light yellow
            'poor': PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'),     # Light pink
            'fail': PatternFill(start_color='FFA07A', end_color='FFA07A', fill_type='solid')      # Light salmon
        }
    
    def generate_class_report(self, class_id: int, term_id: int, output_path: str) -> Tuple[bool, str]:
        """
        Generate comprehensive grade report for a class and term.
        
        Args:
            class_id: Class ID
            term_id: Term ID
            output_path: Output Excel file path
            
        Returns:
            Tuple of (success, message)
        """
        try:
            # Get class and term information
            class_info = self._get_class_info(class_id)
            term_info = self._get_term_info(term_id)
            
            if not class_info or not term_info:
                return False, "Informations classe ou trimestre introuvables"
            
            # Get students and grades
            students_data = self._get_students_grades(class_id, term_id)
            
            if not students_data:
                return False, "Aucune donnée trouvée pour cette classe et ce trimestre"
            
            # Get subjects
            subjects = self._get_subjects_for_class(class_id, term_id)
            
            # Create Excel workbook
            workbook = openpyxl.Workbook()
            
            # Create summary sheet
            self._create_summary_sheet(workbook, class_info, term_info, students_data, subjects)
            
            # Create detailed sheets
            self._create_detailed_sheets(workbook, class_info, term_info, students_data, subjects)
            
            # Create statistics sheet
            self._create_statistics_sheet(workbook, students_data, subjects)
            
            # Save workbook
            workbook.save(output_path)
            
            return True, f"Rapport généré: {len(students_data)} étudiants"
            
        except Exception as e:
            self.logger.error(f"Erreur génération rapport: {e}")
            return False, f"Erreur: {str(e)}"
    
    def _get_class_info(self, class_id: int) -> Optional[Dict]:
        """Get class information."""
        query = QSqlQuery(self.db)
        query.prepare("SELECT name, level, section, academic_year FROM CLASSES WHERE id = ?")
        query.addBindValue(class_id)
        query.exec()
        
        if query.next():
            return {
                'id': class_id,
                'name': query.value(0),
                'level': query.value(1),
                'section': query.value(2),
                'academic_year': query.value(3)
            }
        return None
    
    def _get_term_info(self, term_id: int) -> Optional[Dict]:
        """Get term information."""
        query = QSqlQuery(self.db)
        query.prepare("SELECT name, start_date, end_date FROM TERMS WHERE id = ?")
        query.addBindValue(term_id)
        query.exec()
        
        if query.next():
            return {
                'id': term_id,
                'name': query.value(0),
                'start_date': query.value(1),
                'end_date': query.value(2)
            }
        return None
    
    def _get_students_grades(self, class_id: int, term_id: int) -> List[Dict]:
        """Get students and their grades for the specified class and term."""
        query = QSqlQuery(self.db)
        query.exec(f"""
            SELECT 
                r.id as roster_id,
                r.student_name,
                r.student_id_number,
                s.name as subject_name,
                s.code as subject_code,
                s.coefficient,
                g.grade,
                g.ocr_confidence,
                g.manual_override
            FROM ROSTERS r
            LEFT JOIN GRADES g ON r.id = g.roster_id AND g.term_id = {term_id}
            LEFT JOIN SUBJECTS s ON g.subject_id = s.id
            WHERE r.class_id = {class_id}
            ORDER BY r.student_name, s.name
        """)
        
        # Organize data by student
        students = {}
        while query.next():
            roster_id = query.value(0)
            student_name = query.value(1)
            student_id = query.value(2)
            subject_name = query.value(3)
            subject_code = query.value(4)
            coefficient = query.value(5)
            grade = query.value(6)
            confidence = query.value(7)
            manual_override = query.value(8)
            
            if roster_id not in students:
                students[roster_id] = {
                    'roster_id': roster_id,
                    'name': student_name,
                    'student_id': student_id,
                    'grades': {},
                    'total_points': 0,
                    'total_coefficients': 0,
                    'average': 0
                }
            
            if subject_name and grade is not None:
                students[roster_id]['grades'][subject_code] = {
                    'subject_name': subject_name,
                    'grade': float(grade),
                    'coefficient': float(coefficient) if coefficient else 1.0,
                    'confidence': float(confidence) if confidence else 100.0,
                    'manual_override': bool(manual_override)
                }
        
        # Calculate averages
        for student in students.values():
            total_points = 0
            total_coefficients = 0
            
            for grade_info in student['grades'].values():
                points = grade_info['grade'] * grade_info['coefficient']
                total_points += points
                total_coefficients += grade_info['coefficient']
            
            if total_coefficients > 0:
                student['average'] = total_points / total_coefficients
                student['total_points'] = total_points
                student['total_coefficients'] = total_coefficients
        
        return list(students.values())
    
    def _get_subjects_for_class(self, class_id: int, term_id: int) -> List[Dict]:
        """Get subjects that have grades for this class and term."""
        query = QSqlQuery(self.db)
        query.exec(f"""
            SELECT DISTINCT s.id, s.name, s.code, s.coefficient, s.category
            FROM SUBJECTS s
            JOIN GRADES g ON s.id = g.subject_id
            JOIN ROSTERS r ON g.roster_id = r.id
            WHERE r.class_id = {class_id} AND g.term_id = {term_id}
            ORDER BY s.name
        """)
        
        subjects = []
        while query.next():
            subjects.append({
                'id': query.value(0),
                'name': query.value(1),
                'code': query.value(2),
                'coefficient': query.value(3),
                'category': query.value(4)
            })
        
        return subjects
    
    def _create_summary_sheet(self, workbook: openpyxl.Workbook, class_info: Dict, 
                            term_info: Dict, students_data: List[Dict], subjects: List[Dict]):
        """Create the main summary sheet."""
        ws = workbook.active
        ws.title = "Relevé de Notes"
        
        # Header information
        ws['A1'] = f"RELEVÉ DE NOTES - {class_info['name']}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"Niveau: {class_info['level']} - Section: {class_info['section']}"
        ws['A3'] = f"Année scolaire: {class_info['academic_year']}"
        ws['A4'] = f"Période: {term_info['name']}"
        ws['A5'] = f"Date d'édition: {datetime.now().strftime('%d/%m/%Y')}"
        
        # Create headers
        start_row = 7
        headers = ['N°', 'Nom de l\'étudiant', 'N° Étudiant']
        
        # Add subject headers
        for subject in subjects:
            headers.append(f"{subject['name']}\n(Coef. {subject['coefficient']})")
        
        headers.extend(['Moyenne', 'Rang', 'Observation'])
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.thin_border
            if col > 3 and col <= 3 + len(subjects):
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Sort students by average (descending)
        students_data.sort(key=lambda x: x['average'], reverse=True)
        
        # Write student data
        for idx, student in enumerate(students_data, 1):
            row = start_row + idx
            
            # Basic info
            ws.cell(row=row, column=1, value=idx).border = self.thin_border
            ws.cell(row=row, column=2, value=student['name']).border = self.thin_border
            ws.cell(row=row, column=3, value=student['student_id'] or '').border = self.thin_border
            
            # Grades
            for col, subject in enumerate(subjects, 4):
                subject_code = subject['code']
                if subject_code in student['grades']:
                    grade = student['grades'][subject_code]['grade']
                    cell = ws.cell(row=row, column=col, value=grade)
                    cell.border = self.thin_border
                    cell.alignment = self.center_align
                    cell.font = self.grade_font
                    
                    # Color code based on grade
                    cell.fill = self._get_grade_color(grade)
                else:
                    cell = ws.cell(row=row, column=col, value='--')
                    cell.border = self.thin_border
                    cell.alignment = self.center_align
            
            # Average
            avg_cell = ws.cell(row=row, column=4 + len(subjects), value=round(student['average'], 2))
            avg_cell.border = self.thin_border
            avg_cell.alignment = self.center_align
            avg_cell.font = Font(bold=True, size=11)
            avg_cell.fill = self._get_grade_color(student['average'])
            
            # Rank
            rank_cell = ws.cell(row=row, column=5 + len(subjects), value=idx)
            rank_cell.border = self.thin_border
            rank_cell.alignment = self.center_align
            
            # Observation
            obs_cell = ws.cell(row=row, column=6 + len(subjects), value=self._get_observation(student['average']))
            obs_cell.border = self.thin_border
            obs_cell.alignment = self.left_align
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_detailed_sheets(self, workbook: openpyxl.Workbook, class_info: Dict,
                              term_info: Dict, students_data: List[Dict], subjects: List[Dict]):
        """Create detailed sheets for each subject."""
        for subject in subjects:
            ws = workbook.create_sheet(title=f"Détail {subject['code']}")
            
            # Header
            ws['A1'] = f"DÉTAIL - {subject['name']}"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:F1')
            
            ws['A2'] = f"Classe: {class_info['name']} - {term_info['name']}"
            ws['A3'] = f"Coefficient: {subject['coefficient']}"
            
            # Headers
            headers = ['Rang', 'Nom de l\'étudiant', 'Note', 'Confiance OCR', 'Modification manuelle', 'Observation']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = self.header_font
                cell.border = self.thin_border
                cell.alignment = self.center_align
            
            # Get students with this subject
            subject_students = []
            for student in students_data:
                if subject['code'] in student['grades']:
                    grade_info = student['grades'][subject['code']]
                    subject_students.append({
                        'name': student['name'],
                        'grade': grade_info['grade'],
                        'confidence': grade_info['confidence'],
                        'manual_override': grade_info['manual_override']
                    })
            
            # Sort by grade
            subject_students.sort(key=lambda x: x['grade'], reverse=True)
            
            # Write data
            for idx, student in enumerate(subject_students, 1):
                row = 5 + idx
                
                ws.cell(row=row, column=1, value=idx).border = self.thin_border
                ws.cell(row=row, column=2, value=student['name']).border = self.thin_border
                
                grade_cell = ws.cell(row=row, column=3, value=student['grade'])
                grade_cell.border = self.thin_border
                grade_cell.alignment = self.center_align
                grade_cell.fill = self._get_grade_color(student['grade'])
                
                conf_cell = ws.cell(row=row, column=4, value=f"{student['confidence']:.1f}%")
                conf_cell.border = self.thin_border
                conf_cell.alignment = self.center_align
                
                manual_cell = ws.cell(row=row, column=5, value="Oui" if student['manual_override'] else "Non")
                manual_cell.border = self.thin_border
                manual_cell.alignment = self.center_align
                
                obs_cell = ws.cell(row=row, column=6, value=self._get_observation(student['grade']))
                obs_cell.border = self.thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_statistics_sheet(self, workbook: openpyxl.Workbook, students_data: List[Dict], subjects: List[Dict]):
        """Create statistics sheet."""
        ws = workbook.create_sheet(title="Statistiques")
        
        ws['A1'] = "STATISTIQUES DE CLASSE"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # Overall statistics
        averages = [s['average'] for s in students_data if s['average'] > 0]
        
        if averages:
            ws['A3'] = "Statistiques générales:"
            ws['A3'].font = Font(bold=True)
            
            ws['A4'] = f"Nombre d'étudiants: {len(students_data)}"
            ws['A5'] = f"Moyenne de classe: {sum(averages)/len(averages):.2f}"
            ws['A6'] = f"Note la plus haute: {max(averages):.2f}"
            ws['A7'] = f"Note la plus basse: {min(averages):.2f}"
            
            # Grade distribution
            ws['A9'] = "Répartition des notes:"
            ws['A9'].font = Font(bold=True)
            
            excellent = len([a for a in averages if a >= 16])
            good = len([a for a in averages if 14 <= a < 16])
            average = len([a for a in averages if 12 <= a < 14])
            poor = len([a for a in averages if 10 <= a < 12])
            fail = len([a for a in averages if a < 10])
            
            ws['A10'] = f"Excellent (≥16): {excellent} étudiants ({excellent/len(averages)*100:.1f}%)"
            ws['A11'] = f"Bien (14-16): {good} étudiants ({good/len(averages)*100:.1f}%)"
            ws['A12'] = f"Assez bien (12-14): {average} étudiants ({average/len(averages)*100:.1f}%)"
            ws['A13'] = f"Passable (10-12): {poor} étudiants ({poor/len(averages)*100:.1f}%)"
            ws['A14'] = f"Insuffisant (<10): {fail} étudiants ({fail/len(averages)*100:.1f}%)"
        
        # Subject statistics
        ws['A16'] = "Statistiques par matière:"
        ws['A16'].font = Font(bold=True)
        
        headers = ['Matière', 'Moyenne', 'Note max', 'Note min', 'Étudiants notés']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=17, column=col, value=header)
            cell.font = self.header_font
            cell.border = self.thin_border
        
        for idx, subject in enumerate(subjects):
            row = 18 + idx
            subject_grades = []
            
            for student in students_data:
                if subject['code'] in student['grades']:
                    subject_grades.append(student['grades'][subject['code']]['grade'])
            
            if subject_grades:
                ws.cell(row=row, column=1, value=subject['name']).border = self.thin_border
                ws.cell(row=row, column=2, value=f"{sum(subject_grades)/len(subject_grades):.2f}").border = self.thin_border
                ws.cell(row=row, column=3, value=f"{max(subject_grades):.2f}").border = self.thin_border
                ws.cell(row=row, column=4, value=f"{min(subject_grades):.2f}").border = self.thin_border
                ws.cell(row=row, column=5, value=len(subject_grades)).border = self.thin_border
    
    def _get_grade_color(self, grade: float) -> PatternFill:
        """Get color based on grade."""
        if grade >= 16:
            return self.grade_colors['excellent']
        elif grade >= 14:
            return self.grade_colors['good']
        elif grade >= 12:
            return self.grade_colors['average']
        elif grade >= 10:
            return self.grade_colors['poor']
        else:
            return self.grade_colors['fail']
    
    def _get_observation(self, grade: float) -> str:
        """Get observation based on grade."""
        if grade >= 16:
            return "Excellent"
        elif grade >= 14:
            return "Bien"
        elif grade >= 12:
            return "Assez bien"
        elif grade >= 10:
            return "Passable"
        else:
            return "Insuffisant"
    
    def generate_student_report(self, roster_id: int, term_id: int, output_path: str) -> Tuple[bool, str]:
        """Generate individual student report."""
        try:
            # Get student info
            query = QSqlQuery(self.db)
            query.prepare("""
                SELECT r.student_name, r.student_id_number, c.name as class_name
                FROM ROSTERS r
                JOIN CLASSES c ON r.class_id = c.id
                WHERE r.id = ?
            """)
            query.addBindValue(roster_id)
            query.exec()
            
            if not query.next():
                return False, "Étudiant introuvable"
            
            student_name = query.value(0)
            student_id = query.value(1)
            class_name = query.value(2)
            
            # Get grades
            grades_query = QSqlQuery(self.db)
            grades_query.prepare("""
                SELECT s.name, s.coefficient, g.grade, g.ocr_confidence
                FROM GRADES g
                JOIN SUBJECTS s ON g.subject_id = s.id
                WHERE g.roster_id = ? AND g.term_id = ?
                ORDER BY s.name
            """)
            grades_query.addBindValue(roster_id)
            grades_query.addBindValue(term_id)
            grades_query.exec()
            
            # Create workbook
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "Bulletin individuel"
            
            # Header
            ws['A1'] = "BULLETIN INDIVIDUEL"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:D1')
            
            ws['A3'] = f"Étudiant: {student_name}"
            ws['A4'] = f"N° Étudiant: {student_id or 'N/A'}"
            ws['A5'] = f"Classe: {class_name}"
            
            # Grades table
            headers = ['Matière', 'Coefficient', 'Note', 'Observation']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col, value=header)
                cell.font = self.header_font
                cell.border = self.thin_border
            
            total_points = 0
            total_coefficients = 0
            row = 8
            
            while grades_query.next():
                subject_name = grades_query.value(0)
                coefficient = grades_query.value(1)
                grade = grades_query.value(2)
                
                ws.cell(row=row, column=1, value=subject_name).border = self.thin_border
                ws.cell(row=row, column=2, value=coefficient).border = self.thin_border
                
                grade_cell = ws.cell(row=row, column=3, value=grade)
                grade_cell.border = self.thin_border
                grade_cell.fill = self._get_grade_color(grade)
                
                ws.cell(row=row, column=4, value=self._get_observation(grade)).border = self.thin_border
                
                total_points += grade * coefficient
                total_coefficients += coefficient
                row += 1
            
            # Average
            if total_coefficients > 0:
                average = total_points / total_coefficients
                avg_row = row + 1
                ws.cell(row=avg_row, column=1, value="MOYENNE GÉNÉRALE").font = Font(bold=True)
                avg_cell = ws.cell(row=avg_row, column=3, value=f"{average:.2f}")
                avg_cell.font = Font(bold=True)
                avg_cell.fill = self._get_grade_color(average)
                
                ws.cell(row=avg_row, column=4, value=self._get_observation(average)).font = Font(bold=True)
            
            workbook.save(output_path)
            return True, f"Bulletin généré pour {student_name}"
            
        except Exception as e:
            self.logger.error(f"Erreur génération bulletin: {e}")
            return False, f"Erreur: {str(e)}"


def create_grade_template(output_path: str) -> bool:
    """Create a template Excel file for grade entry."""
    try:
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Modèle de Notes"
        
        # Headers
        ws['A1'] = "NOM DE L'ÉTUDIANT"
        ws['B1'] = "MATHÉMATIQUES"
        ws['C1'] = "FRANÇAIS"
        ws['D1'] = "ARABE"
        ws['E1'] = "SCIENCES"
        ws['F1'] = "HISTOIRE-GÉO"
        
        # Sample data
        students = [
            "Ahmed Ben Ali",
            "Fatima El Mansouri", 
            "Mohammed Karimi",
            "Aicha Benjelloun",
            "Youssef Tazi"
        ]
        
        for idx, student in enumerate(students, 2):
            ws.cell(row=idx, column=1, value=student)
        
        # Formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        workbook.save(output_path)
        return True
        
    except Exception as e:
        logging.error(f"Erreur création template: {e}")
        return False
